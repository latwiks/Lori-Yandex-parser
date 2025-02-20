import csv
import os
import sqlite3
from openpyxl import Workbook, load_workbook
import ast


async def isint(x):
    try:
        a = float(x)
        b = int(a)
    except (TypeError, ValueError):
        return False
    else:
        return a == b


async def excel_2(db, url, limiter, uid, criminal_id, sortby, isphoto):
    conn = sqlite3.connect(f'databases/{db}')
    c = conn.cursor()
    link = ast.literal_eval(c.execute('select link from user where id=?', (uid,)).fetchone()[0])
    times = ast.literal_eval(c.execute('select times from user where id=?', (uid,)).fetchone()[0])
    criminal_links = ast.literal_eval(
        c.execute('select criminal_links from criminals where id=?', (criminal_id,)).fetchone()[0])
    criminal_descriptions = ast.literal_eval(
        c.execute('select criminal_descriptions from criminals where id=?', (criminal_id,)).fetchone()[0])
    pages = ast.literal_eval(c.execute('select pages from user where id=?', (uid,)).fetchone()[0])
    conn.close()
    wb = Workbook()
    ws = wb.active
    data = [['№ страницы', '№ картинки', 'Профиль', 'Ссылка на картинку', 'Дата съемки', '№ сайта', 'Ссылки на сайты',
             'Описание сайта']]
    for row in data:
        ws.append(row)
    for i in range(len(link)):
        for j in range(len(link[i])):
            if len(criminal_links[i][j]) >= 20:
                if await isint(limiter) or limiter == '':
                    if limiter == '':
                        ws.append([pages[i][j], j + 1, url, link[i][j], times[i][j], 1, 'Более 20 результатов',
                                   'Точных совпадений не выявлено'])
                        continue
                    elif int(limiter) <= 0:
                        ws.append([pages[i][j], j + 1, url, link[i][j], times[i][j], 1, 'Более 20 результатов',
                                   'Точных совпадений не выявлено'])
                        continue
                    else:
                        data1 = []
                        ij = 1
                        for ii in range(len(criminal_links[i][j])):
                            if ij + 1 > int(limiter) + 1:
                                break
                            data1.append(
                                [pages[i][j], j + 1, url, link[i][j], times[i][j], ij, criminal_links[i][j][ii],
                                 criminal_descriptions[i][j][ii][0]])
                            ij += 1
                        for row in data1:
                            ws.append(row)
                        continue

            else:
                data1 = []
                ij = 1
                for ii in range(len(criminal_links[i][j])):
                    data1.append([pages[i][j], j + 1, url, link[i][j], times[i][j], ij, criminal_links[i][j][ii],
                                  criminal_descriptions[i][j][ii][0]])
                    ij += 1
                for row in data1:
                    ws.append(row)
    if isphoto:
        a = '_'
    elif sortby == '?':
        a = '_by_order_'
    elif sortby == '?sort=date':
        a = '_by_date_'
    elif sortby == '?sort=sales':
        a = '_by_sales_'
    elif sortby == '?sort=random':
        a = '_by_random_'
    wb.save(f"excel_files/{url[16:]}{a}{uid}_{criminal_id}.xlsx")
    wb.close()
    return f'excel_files/{url[16:]}{a}{uid}_{criminal_id}.xlsx'


async def remove_duplicates_xlsx(db, uid, criminal_id, url, sortby, isphoto):
    if isphoto:
        a = '_'
    elif sortby == '?':
        a = '_by_order_'
    elif sortby == '?sort=date':
        a = '_by_date_'
    elif sortby == '?sort=sales':
        a = '_by_sales_'
    elif sortby == '?sort=random':
        a = '_by_random_'
    unique_links = set()
    wb = load_workbook(f'excel_files/{url}{a}{uid}_{criminal_id}.xlsx')
    ws = wb.active
    new_rows = [ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        link = row[6]
        if link not in unique_links or link == 'Более 20 результатов':
            unique_links.add(link)
            new_rows.append(row)

    ws.delete_rows(1, ws.max_row)
    for row in new_rows:
        ws.append(row)

    wb.save(f'excel_files/{url}{a}{uid}_{criminal_id}_unique.xlsx')
    wb.close()
    return f'excel_files/{url}{a}{uid}_{criminal_id}_unique.xlsx'


async def excel_to_csv_2(db, url, limiter, uid, criminal_id, sortby, isphoto):
    conn = sqlite3.connect(f'databases/{db}')
    c = conn.cursor()

    link = ast.literal_eval(c.execute('select link from user where id=?', (uid,)).fetchone()[0])
    times = ast.literal_eval(c.execute('select times from user where id=?', (uid,)).fetchone()[0])
    criminal_links = ast.literal_eval(
        c.execute('select criminal_links from criminals where id=?', (criminal_id,)).fetchone()[0])
    criminal_descriptions = ast.literal_eval(
        c.execute('select criminal_descriptions from criminals where id=?', (criminal_id,)).fetchone()[0])
    pages = ast.literal_eval(c.execute('select pages from user where id=?', (uid,)).fetchone()[0])
    conn.close()

    if isphoto:
        a = '_'
    elif sortby == '?':
        a = '_by_order_'
    elif sortby == '?sort=date':
        a = '_by_date_'
    elif sortby == '?sort=sales':
        a = '_by_sales_'
    elif sortby == '?sort=random':
        a = '_by_random_'

    with open(f'csv_files/{url[16:]}{a}{uid}_{criminal_id}.csv', 'w', newline='', encoding='utf-16') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow(
            ['№ страницы', '№ картинки', 'Профиль', 'Ссылка на картинку', 'Дата съемки', '№ сайта', 'Ссылки на сайты',
             'Описание сайта'])
        for i in range(len(link)):
            for j in range(len(link[i])):
                if len(criminal_links[i][j]) >= 20:
                    if await isint(limiter) or limiter == '':
                        if limiter == '':
                            writer.writerow(
                                [pages[i][j], j + 1, url, link[i][j], times[i][j], 1, 'Более 20 результатов',
                                 'Точных совпадений не выявлено'])
                        elif int(limiter) <= 0:
                            writer.writerow(
                                [pages[i][j], j + 1, url, link[i][j], times[i][j], 1, 'Более 20 результатов',
                                 'Точных совпадений не выявлено'])
                        else:
                            for ii in range(len(criminal_links[j])):
                                if ii + 1 > int(limiter):
                                    break
                                writer.writerow(
                                    [pages[i][j], j + 1, url, link[i][j], times[i][j], ii + 1,
                                     criminal_links[i][j][ii],
                                     criminal_descriptions[i][j][ii][0]])
                else:
                    for ii in range(len(criminal_links[j])):
                        writer.writerow(
                            [pages[i][j], j + 1, url, link[i][j], times[i][j], ii + 1, criminal_links[i][j][ii],
                             criminal_descriptions[i][j][ii][0]])
    return f'csv_files/{url[16:]}{a}{uid}_{criminal_id}.csv'


async def remove_duplicates_csv(db, uid, criminal_id, url, sortby, isphoto):
    if isphoto:
        a = '_'
    elif sortby == '?':
        a = '_by_order_'
    elif sortby == '?sort=date':
        a = '_by_date_'
    elif sortby == '?sort=sales':
        a = '_by_sales_'
    elif sortby == '?sort=random':
        a = '_by_random_'
    unique_links = set()
    new_rows = []
    with open(f'csv_files/{url}{a}{uid}_{criminal_id}.csv', 'r', encoding='utf-16') as file:
        reader = csv.reader(file, delimiter=';')
        headers = next(reader)
        for row in reader:
            link = row[6]
            if link not in unique_links or link == 'Более 20 результатов':
                unique_links.add(link)
                new_rows.append(row)

    with open(f'csv_files/{url}{a}{uid}_{criminal_id}_unique.csv', 'w', newline='', encoding='utf-16') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow(headers)
        writer.writerows(new_rows)
    os.remove(f'csv_files/{url}{a}{uid}_{criminal_id}.csv')
    return f'csv_files/{url}{a}{uid}_{criminal_id}_unique.csv'
